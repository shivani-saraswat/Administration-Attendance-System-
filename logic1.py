#logic.py
import cv2
import pytz
import sqlite3
import numpy as np
# import face_recognition
from datetime import datetime
import pandas as pd
import holidays
from models1 import *
from fastapi import HTTPException, status
from io import BytesIO
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def to_binary(image):
    _, buffer = cv2.imencode('.jpg', image)
    return buffer.tobytes()

def init_db():
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute('PRAGMA foreign_keys = ON')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS faces (
            SrNo INTEGER,
            image BLOB,
            image_en,
            FOREIGN KEY(SrNo) REFERENCES users(SrNo)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            SrNo INTEGER,
            date TEXT,
            in_time TEXT,
            out_time TEXT,
            status TEXT DEFAULT 'Absent',
            FOREIGN KEY(SrNo) REFERENCES users(SrNo)
        )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
    SrNo INTEGER PRIMARY KEY AUTOINCREMENT,
    Reporting_to TEXT,
    Joining_date TEXT,
    Location TEXT,
    Email TEXT,
    Emp_id TEXT,
    Name TEXT,
    Department TEXT,
    Password TEXT,
    Role TEXT
    )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS token_blacklist (
            token TEXT PRIMARY KEY,
            expiry TIMESTAMP
        )

    ''')
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS deleted_users (
        SrNo INTEGER,
        Reporting_to TEXT,
        Joining_date TEXT,
        Location TEXT,
        Email TEXT,
        Emp_id TEXT,
        Name TEXT,
        Department TEXT
        )
    """)
    
    conn.commit()
    conn.close()

def id_exists(Emp_id):
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute("SELECT Emp_id FROM users WHERE Emp_id = ? COLLATE NOCASE", (Emp_id,))
    exists = cursor.fetchone() is not None
    conn.close()
    return exists

def save_face_to_db(emp_id, face_img, image_en):
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute("Select SrNo from users where Emp_id = ? COLLATE NOCASE", (emp_id,))
    row = cursor.fetchone()
    srno = int(row[0])
    cursor.execute("""
        INSERT INTO faces (SrNo, image, image_en)
        VALUES (?, ?, ?)
    """, (srno,to_binary(face_img), image_en))
    conn.commit()
    conn.close()

def get_face_encodings_from_db():
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute("SELECT u.SrNo, u.Emp_id, u.Name, u.Department, u.Reporting_to, u.Location, u.Joining_date, f.image_en FROM users u join faces f on u.SrNo = f.SrNo")
    rows = cursor.fetchall()
    conn.close()

    encodings = []
    metadata = []
    for SrNo, Emp_id, Name, Department, Reporting_to, Location, Joining_date, img_encs in rows:
        if img_encs:
            encoding_array = np.frombuffer(img_encs, dtype=np.float64)  # dtype must match what you saved
            encodings.append(encoding_array)
            metadata.append((SrNo, Emp_id, Name, Department, Reporting_to, Location, Joining_date))
    return encodings, metadata

def is_off_day():
    tz = pytz.timezone('Asia/Kolkata')
    now = datetime.now(tz)
    today = now.date()
    # india_holidays = holidays.India(year=today.year)
    india_holidays = holidays.country_holidays('IN')
    # return today in india_holidays or today.weekday() >= 5
    return today in india_holidays or today.weekday() >= 5 
    #   # 5 = Saturday, 6 = Sunday

def compare_is_off_day(date_str):
    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    
    # Create the India holiday calendar for a range of years if needed
    india_holidays = holidays.country_holidays('IN')  # No 'year' parameter
    
    # Return True if it's a weekend or a holiday
    return date_obj in india_holidays or date_obj.weekday() >= 5 

# --- Modified Attendance Marking ---

def mark_attendance(SrNo):
    print(SrNo)
    tz = pytz.timezone('Asia/Kolkata')
    now = datetime.now(tz)
    # now = datetime.now()
    # custom_date = datetime(2025, 7, 6, now.hour, now.minute, now.second, now.microsecond)
    # print(custom_date)
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()

    # Check if already marked
    cursor.execute("SELECT * FROM attendance WHERE SrNo = ? AND date = ?", (SrNo, current_date))
    record = cursor.fetchone()

    if is_off_day():
        if not record:
            # Mark as Present with default times on holiday/weekend
            cursor.execute("""
                INSERT INTO attendance (SrNo, date, in_time, out_time, status)
                VALUES (?, ?, ?, ?, 'Present')
            """, (SrNo, current_date, '00:00:00', '00:00:00'))
            conn.commit()
            conn.close()
            return "Marked Present (Weekend/Holiday)"
        else:
            conn.close()
            return "Already marked for holiday/weekend."

    # Weekday logic
    if not record:
        cursor.execute("""
            INSERT INTO attendance (SrNo, date, in_time, status)
            VALUES (?, ?, ?, 'Present')
        """, (SrNo, current_date, current_time))
    else:
        cursor.execute("""
            UPDATE attendance SET out_time = ? WHERE SrNo = ? AND date = ?
        """, (current_time, SrNo, current_date))

    conn.commit()
    conn.close()
    return "Attendance marked successfully."

# def testing_holiday(date):
#     if not compare_is_off_day(date_str=date):
#         return {"message": "Today is not off"}
#     return {"message": "Today is holiday"}


def auto_mark_all_present_on_off_days():
    tz = pytz.timezone('Asia/Kolkata')
    if not is_off_day():
        return "Today is a working day. No auto-marking."

    conn = sqlite3.connect("faces.db")
    cursor = conn.cursor()
    today = datetime.now(tz).strftime("%Y-%m-%d")

    # Get all employee SrNo
    cursor.execute("SELECT SrNo FROM users")
    srnos = [row[0] for row in cursor.fetchall()]

    for srno in srnos:
        # Check if already marked
        cursor.execute("SELECT 1 FROM attendance WHERE SrNo = ? AND date = ?", (srno, today))
        already_marked = cursor.fetchone()
        if not already_marked:
            cursor.execute("""
                INSERT INTO attendance (SrNo, date, in_time, out_time, status)
                VALUES (?, ?, ?, ?, 'Present')
            """, (srno, today, '00:00:00', '00:00:00'))

    conn.commit()
    conn.close()
    return "Auto-marked present for all on holiday/weekend."


def update_user_password(email, password):
    conn = sqlite3.connect("faces.db")
    cursor = conn.cursor()
    try:
        cursor.execute(f"Update users set Password=? where Email=?",(password, email))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        conn.rollback()
        conn.close()
        return False


# def delete_face_by_id(Emp_id):
#     conn = sqlite3.connect('faces.db')
#     cursor = conn.cursor()
#     cursor.execute("DELETE FROM users WHERE Emp_id = ?", (Emp_id,))
#     conn.commit()
#     deleted_count = cursor.rowcount
#     conn.close()
#     return deleted_count

def delete_face_by_id(Emp_id):
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()

    # Get user data before deleting
    cursor.execute("""
        SELECT SrNo, Reporting_to, Joining_date, Location, Email, Emp_id, Name, Department 
        FROM users WHERE Emp_id = ? COLLATE NOCASE
    """, (Emp_id,))
    user_data = cursor.fetchone()

    if not user_data:
        conn.close()
        return 0

    # Insert into deleted_users table
    cursor.execute("""
        INSERT INTO deleted_users (SrNo, Reporting_to, Joining_date, Location, Email, Emp_id, Name, Department)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, user_data)

    # Update the users table: clear data but keep SrNo and Emp_id
    cursor.execute("""
        UPDATE users
        SET Reporting_to = NULL,
            Joining_date = NULL,
            Location = NULL,
            Email = NULL,
            Name = NULL,
            Department = NULL,
            Password = NULL,
            Role = NULL
        WHERE Emp_id = ? COLLATE NOCASE
    """, (Emp_id,))

    conn.commit()
    deleted_count = cursor.rowcount
    conn.close()
    return deleted_count


def delete_face(emp_id):
    count = delete_face_by_id(emp_id)
    return f"Deleted {count} record(s)" if count else "No matching record found."

# def get_saved_faces():
#     conn = sqlite3.connect('faces.db')
#     cursor = conn.cursor()
#     cursor.execute("SELECT u.SrNo, u.Name, u.Department, f.image FROM faces f join users u on f.Srno = u.SrNo")
#     rows = cursor.fetchall()
#     conn.close()
#     return rows

def get_saved_faces():
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute("""
        SELECT f.SrNo, u.Name, u.Department, f.image
        FROM users u
        JOIN faces f ON u.SrNo = f.SrNo
        WHERE u.Name IS NOT NULL
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_registered_faces():
    conn = sqlite3.connect("faces.db")
    cursor = conn.cursor()
    cursor.execute("SELECT u.Emp_id, u.Name, u.Department, u.Reporting_to, u.Location, u.Joining_date FROM users u")
    rows = cursor.fetchall()
    conn.close()
    return [
        {
            "Emp_id": emp_id,
            "Name": name,
            "Department": dept,
            "Reporting_to": reporting,
            "Location": loc,
            "Joining_date": join
        } for emp_id, name, dept, reporting, loc, join in rows
    ]

# def export_data_to_excel(purpose, file_path='filtered_attendance.xlsx'):
#     conn = sqlite3.connect("faces.db")
#     df = pd.read_sql_query("""
#         SELECT 
#             u.Emp_id,
#             u.Name,
#             u.Department,
#             u.Reporting_to,
#             u.Location,
#             u.Joining_date,
#             a.date,
#             a.in_time,
#             a.out_time
#         FROM attendance a
#         JOIN users u ON a.SrNo = u.SrNo
#         ORDER BY a.date DESC
#     """, conn)
#     conn.close()
#     def compute_working_hours(row):
#         try:
#             if pd.notna(row["in_time"]) and pd.notna(row["out_time"]):
#                 in_time = datetime.strptime(row["in_time"], "%H:%M:%S")
#                 out_time = datetime.strptime(row["out_time"], "%H:%M:%S")
#                 return round((out_time - in_time).total_seconds() / 3600, 2)
#         except:
#             return None

#     df["Working_Hours"] = df.apply(compute_working_hours, axis=1)

#     def compute_status(hours):
#         if pd.isna(hours): return "Absent"
#         if hours >= 9: return "Present"
#         elif hours >= 4: return "Half Day"
#         return "Absent"

#     # print(df.head())
#     df["status"] = df["Working_Hours"].apply(compute_status)
#     df["Working_Hours"] = df["Working_Hours"].astype(str)
#     # print(df.head())
#     # if date_filter:
#     #     df = df[df["date"] == date_filter]
#     # if status_filter:
#     #     df = df[df["Status"] == status_filter]
#     if purpose=="download":
#     # Create a BytesIO object to store the Excel file in memory
#         output = BytesIO()
#         df.to_excel(output, index=False)
#         output.seek(0)
#         return output.getvalue()
#     elif purpose=="view":
#         # print(df.head())
#         return df.to_dict(orient="records")


def export_data_to_excel(purpose, filter_date=None, file_path='filtered_attendance.xlsx'):
    conn = sqlite3.connect("faces.db")
    df = pd.read_sql_query("""
        SELECT 
            u.Emp_id,
            u.Name,
            u.Reporting_to,
            u.Location,
            u.Department,
            u.Joining_date,
            a.date,
            a.in_time,
            a.out_time
        FROM attendance a
        JOIN users u ON a.SrNo = u.SrNo
        ORDER BY a.date DESC
    """, conn)
    conn.close()

    # Convert date columns to datetime
    df["date"] = pd.to_datetime(df["date"])
    df["Joining_date"] = pd.to_datetime(df["Joining_date"], errors='coerce')

    # Get all unique months in the data
    df['month'] = df['date'].dt.to_period('M')
    months = df['month'].dropna().unique()

    # Get all employees
    employees = df[['Emp_id', 'Name', 'Reporting_to', 'Location', 'Joining_date']].drop_duplicates()

    if purpose == "download":
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        for month in months:
            month_str = str(month)
            month_df = df[df['month'] == month]
            # Get all days in this month
            year, mon = map(int, month_str.split('-'))
            num_days = calendar.monthrange(year, mon)[1]
            days = [datetime(year, mon, d+1).strftime('%d-%b-%y') for d in range(num_days)]
            # Prepare data for this sheet
            data = []
            for _, emp in employees.iterrows():
                row = [emp['Emp_id'], emp['Name'], emp['Reporting_to'], emp['Location'], emp['Joining_date'].strftime('%d-%b-%y') if pd.notna(emp['Joining_date']) else '']
                emp_att = month_df[month_df['Emp_id'] == emp['Emp_id']]
                att_map = {d.strftime('%d-%b-%y'): '' for d in pd.date_range(f'{year}-{mon:02d}-01', periods=num_days)}
                for _, att in emp_att.iterrows():
                    day_str = att['date'].strftime('%d-%b-%y')
                    if pd.notna(att['in_time']) and pd.notna(att['out_time']):
                        in_time = datetime.strptime(att['in_time'], "%H:%M:%S")
                        out_time = datetime.strptime(att['out_time'], "%H:%M:%S")
                        hours = (out_time - in_time).total_seconds() / 3600
                        if hours >= 9:
                            att_map[day_str] = 'P'
                        elif hours >= 4:
                            att_map[day_str] = 'HD'
                        else:
                            att_map[day_str] = 'A'
                    else:
                        att_map[day_str] = 'A'
                row.extend([att_map[d] for d in days])
                data.append(row)
            # Create sheet
            ws = wb.create_sheet(title=month.strftime('%b-%Y'))
            # Header
            headers = ['Employee Id', 'Full Name', 'Reporting To', 'Location', 'Joining Date'] + days
            ws.append(headers)
            # Style header
            header_fill = PatternFill(start_color='B7E1CD', end_color='B7E1CD', fill_type='solid')
            date_fill = PatternFill(start_color='00CFEF', end_color='00CFEF', fill_type='solid')
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                if col <= 5:
                    cell.fill = header_fill
                else:
                    cell.fill = date_fill
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # Data rows
            for row in data:
                ws.append(row)
            # Style data
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in r:
                    c.alignment = Alignment(horizontal='center')
                    c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # Set column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    elif purpose == "view":
        # Restore previous logic for view
        def compute_working_hours(row):
            try:
                if pd.notna(row["in_time"]) and pd.notna(row["out_time"]):
                    in_time = datetime.strptime(row["in_time"], "%H:%M:%S")
                    out_time = datetime.strptime(row["out_time"], "%H:%M:%S")
                    return round((out_time - in_time).total_seconds() / 3600, 2)
            except:
                return None
        df["Working_Hours"] = df.apply(compute_working_hours, axis=1)
        def compute_status(hours):
            if pd.isna(hours): return "Absent"
            if hours >= 9: return "Present"
            elif hours >= 4: return "Half Day"
            return "Absent"
        df["status"] = df["Working_Hours"].apply(compute_status)
        df["Working_Hours"] = df["Working_Hours"].astype(str)
        # Apply date filter if provided
        if filter_date:
            df = df[df["date"] == filter_date]
        # Convert datetime columns to string for API response
        df["Joining_date"] = df["Joining_date"].dt.strftime("%Y-%m-%d")
        df["date"] = df["date"].dt.strftime("%Y-%m-%d")
        return df.to_dict(orient="records")

def check_blacklist_status(token):
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM token_blacklist WHERE token = ?", (token,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=401, detail="Token revoked")
    conn.close()

# def search_attendance_by_name(name_query):
#     conn = sqlite3.connect('faces.db')
#     df = pd.read_sql_query("""
#         SELECT 
#             f.Emp_id,
#             f.Name,
#             f.Department,
#             f.Reporting_to,
#             f.Location,
#             f.Joining_date,
#             a.date,
#             a.in_time,
#             a.out_time
#         FROM attendance a
#         JOIN faces f ON a.SrNo = f.SrNo
#         WHERE f.Name LIKE ?
#     """, conn, params=(f'%{name_query.strip()}%',))
#     conn.close()

#     def compute_working_hours(row):
#         try:
#             if pd.notna(row["in_time"]) and pd.notna(row["out_time"]):
#                 in_time = datetime.strptime(row["in_time"], "%H:%M:%S")
#                 out_time = datetime.strptime(row["out_time"], "%H:%M:%S")
#                 return str(round((out_time - in_time).total_seconds() / 3600, 2))
#         except:
#             return None

#     df["Working Hours"] = df.apply(compute_working_hours, axis=1)
#     df["Status"] = df["Working Hours"].apply(lambda h: "Present" if h >= 9 else "Half Day" if h >= 4 else "Absent")

#     return df

# def capture_and_register(Emp_id, Name, Reporting_to, Location, Department, Joining_date):
#     if id_exists(Emp_id):
#         return "Employee ID already exists."

#     cap = cv2.VideoCapture(0)
#     ret, frame = cap.read()
#     cap.release()

#     if not ret:
#         return "Failed to access webcam."

#     save_face_to_db(Emp_id, Name, Department, Reporting_to, Location, Joining_date, frame)
#     return f"{Name} registered successfully!"

# def search_attendance_records(name="", date="", status=""):
#     conn = sqlite3.connect("faces.db")
#     query = """
#         SELECT 
#             f.Emp_id,
#             f.Name,
#             f.Department,
#             f.Reporting_to,
#             f.Location,
#             f.Joining_date,
#             a.date,
#             a.in_time,
#             a.out_time
#         FROM attendance a
#         JOIN faces f ON a.SrNo = f.SrNo
#         WHERE 1=1
#     """
#     params = []

#     if name:
#         query += " AND f.Name LIKE ?"
#         params.append(f"%{name}%")
#     if date:
#         query += " AND a.date = ?"
#         params.append(date)

#     df = pd.read_sql_query(query, conn, params=params)
#     conn.close()


def search_attendance_records(name=""):
    conn = sqlite3.connect("faces.db")
    cur = conn.cursor()
    print(name)
    pattern = f"{name}%"
    print(pattern)
    cur.execute(
        """
        SELECT 
            f.Emp_id,
            f.Name,
            f.Department,
            f.Reporting_to,
            f.Location,
            f.Joining_date,
            a.date,
            a.in_time,
            a.out_time
        FROM attendance a
        JOIN users f ON a.SrNo = f.SrNo
        WHERE f.Name LIKE ?
        """,
        (pattern,),
    )
    
    rows = cur.fetchall()
    for row in rows:
        print(row)

    columns = [
        "Emp_id", "Name", "Department", "Reporting_to", "Location",
        "Joining_date", "date", "in_time", "out_time"
    ]
    
    df = pd.DataFrame(rows, columns=columns)
    
    conn.close()
    def compute_working_hours(row):
        try:
            if pd.notna(row["in_time"]) and pd.notna(row["out_time"]):
                in_time = datetime.strptime(row["in_time"], "%H:%M:%S")
                out_time = datetime.strptime(row["out_time"], "%H:%M:%S")
                return round((out_time - in_time).total_seconds() / 3600, 2)
        except:
            return None

    df["Working_Hours"] = df.apply(compute_working_hours, axis=1)

    def compute_status(hours):
        if pd.isna(hours): return "Absent"
        if hours >= 9: return "Present"
        elif hours >= 4: return "Half Day"
        return "Absent"

    df["status"] = df["Working_Hours"].apply(compute_status)
    df["Working_Hours"] = df["Working_Hours"].astype(str)

    print(df.head())
    return df.to_dict(orient="records")

def get_attendance_stats():
    conn = sqlite3.connect("faces.db")
    total_employees = pd.read_sql_query("SELECT COUNT(*) as count FROM users", conn)["count"][0]

    today = datetime.now().strftime("%Y-%m-%d")
    df = pd.read_sql_query("SELECT in_time, out_time FROM attendance WHERE date = ?", conn, params=(today,))
    conn.close()

    def compute_hours(row):
        try:
            if pd.notna(row["in_time"]) and pd.notna(row["out_time"]):
                in_time = datetime.strptime(row["in_time"], "%H:%M:%S")
                out_time = datetime.strptime(row["out_time"], "%H:%M:%S")
                return round((out_time - in_time).total_seconds() / 3600, 2)
        except:
            return None

    df["Hours"] = df.apply(compute_hours, axis=1)

    present = df["Hours"].apply(lambda h: h >= 9 if pd.notna(h) else False).sum()
    half_day = df["Hours"].apply(lambda h: 4 <= h < 9 if pd.notna(h) else False).sum()
    absent = total_employees - (present + half_day)

    return {
        "total": total_employees,
        "present": present,
        "half_day": half_day,
        "absent": absent
    }

def get_incremented_empId():
    conn = sqlite3.connect("faces.db")
    cursor = conn.cursor()
    try:
        cursor.execute("select Emp_id from users ORDER BY SrNo DESC limit 1")
        row = cursor.fetchone()
        
        if row:
            last_id = row[0]
            num = int(last_id.split("-")[1])
            new_num=num + 1
        else: 
            new_num = 1
        
        new_empid = f"WM-{new_num:02d}"
        return new_empid
    except Exception as e:
        print(f"Error Occured: {e}")


def insert_into_userDB(reporting_to, joining_date, location, email, empId, name, department, password, role):
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO users (Reporting_to, Joining_date, Location, Email, Emp_id, Name, Department, Password, Role) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (reporting_to, joining_date, location, email, empId, name, department, password, role)
    )   
    conn.commit()
    conn.close()
    return {"message": "User registered successfully"}


def get_user_from_db(emp_id: Optional[str] = None, email: Optional[str] = None, query: Optional[list[str]] = None):
    if not query:
        raise ValueError("Query fields list must be provided")
    # print(emp_id)
    # print(query)
    conn = sqlite3.connect('faces.db')
    cursor = conn.cursor()
    option = {
        "emp_id": "Emp_id",
        "email": "Email"
    }

    if emp_id:
        # print("I am here")
        cursor.execute(f"SELECT {', '.join(query)} FROM users WHERE {option['emp_id']} = ? COLLATE NOCASE", (emp_id,))
    elif email:
        cursor.execute(f"SELECT {', '.join(query)} FROM users WHERE {option['email']} = ?", (email,))
    else:
        raise ValueError("Either emp_id or email must be provided")

    row = cursor.fetchone()
    cols = [desc[0] for desc in cursor.description]
    # print(cols)
    conn.close()

    if row:
        # map column names to values
        data = dict(zip(cols, row))
        # print(data)
        # print(UserInDB(**data))
        return UserInDB(**data)
    return None    

if __name__ == "__main__":
    init_db()
